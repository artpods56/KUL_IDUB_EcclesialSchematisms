import csv
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from enum import StrEnum
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Annotated, Literal, Self, cast, final, override

from openpyxl import load_workbook
from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    StrictBool,
    StrictFloat,
    StrictInt,
    StrictStr,
    field_validator,
    model_validator,
)
from rapidfuzz import fuzz
from unidecode import unidecode

from grafy_core.artifacts import NodeConfig, NodeInput, NodeOutput
from grafy_core.domain.plugin_capabilities import PluginRuntimeCapability
from grafy_core.nodes import InPort, Node, NodeExecutionContext, OutPort
from grafy_core.plugins import NodeCachePolicy, NodeStagedUploadInput
from grafy_core.ports.staged_uploads import StagedUploadUnitOfWorkPort
from grafy_core.staged_upload_paths import resolve_persisted_staged_upload_path
from grafy_core.table_contracts import (
    TABLE_DATA,
    Table,
    TableColumn,
    TableValue,
    TableValueType,
)

from grafy_plugin_table.declaration import TABLES


class TableFileImportError(RuntimeError):
    pass


class TableFileUploadItem(BaseModel):
    model_config = ConfigDict(extra="forbid")

    upload_key: StrictStr = Field(min_length=1)
    filename: StrictStr = Field(min_length=1, max_length=1_024)
    byte_size: StrictInt = Field(ge=1)


class TableFileImportConfig(NodeConfig):
    uploads: list[TableFileUploadItem] = Field(
        min_length=1,
        max_length=1,
        description="One staged CSV or XLSX file.",
    )
    sheet_name: StrictStr | None = Field(
        default=None,
        min_length=1,
        max_length=255,
        description="XLSX worksheet name. Leave empty to use the active sheet.",
    )
    header_row: StrictInt = Field(
        default=1,
        ge=1,
        description="One-based row containing column titles.",
    )
    delimiter: StrictStr | None = Field(
        default=None,
        min_length=1,
        max_length=1,
        description="CSV delimiter. Leave empty to detect it from the file.",
    )
    skip_empty_rows: StrictBool = True

    @field_validator("sheet_name")
    @classmethod
    def validate_sheet_name(cls, value: str | None) -> str | None:
        if value is not None and value != value.strip():
            raise ValueError("sheet_name must not have surrounding whitespace")
        return value


class TableFileImportInput(NodeInput):
    pass


class TableFileImportOutput(NodeOutput):
    table: Annotated[
        Table,
        OutPort(TABLE_DATA),
        Field(description="Table imported from the selected worksheet or CSV file."),
    ]


def _table_value(value: object) -> TableValue:
    if value is None or isinstance(value, str | bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)


def _inferred_table_value_type(values: list[TableValue]) -> TableValueType:
    observed: set[TableValueType] = set()
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            observed.add(TableValueType.BOOLEAN)
        elif isinstance(value, int):
            observed.add(TableValueType.INTEGER)
        elif isinstance(value, float):
            observed.add(TableValueType.NUMBER)
        elif isinstance(value, str):
            observed.add(TableValueType.TEXT)
        else:
            observed.add(TableValueType.JSON)
    if not observed:
        return TableValueType.UNKNOWN
    if observed <= {TableValueType.INTEGER, TableValueType.NUMBER}:
        return (
            TableValueType.INTEGER
            if observed == {TableValueType.INTEGER}
            else TableValueType.NUMBER
        )
    return next(iter(observed)) if len(observed) == 1 else TableValueType.MIXED


def _table_from_matrix(
    matrix: list[list[object]],
    *,
    header_row: int,
    skip_empty_rows: bool,
) -> Table:
    header_index = header_row - 1
    if header_index >= len(matrix):
        raise TableFileImportError(
            f"Header row {header_row} is outside the imported file"
        )
    data_rows = matrix[header_index + 1 :]
    column_count = max(
        [len(matrix[header_index]), *(len(row) for row in data_rows)],
        default=0,
    )
    if column_count == 0:
        raise TableFileImportError("The imported file does not contain any columns")

    header = matrix[header_index]
    columns = [
        TableColumn(
            id=f"column_{column_index + 1}",
            title=(
                str(header[column_index]).strip()
                if column_index < len(header)
                and header[column_index] is not None
                and str(header[column_index]).strip()
                else f"Column {column_index + 1}"
            ),
        )
        for column_index in range(column_count)
    ]
    rows: list[dict[str, TableValue]] = []
    for raw_row in data_rows:
        values = [
            _table_value(raw_row[column_index] if column_index < len(raw_row) else None)
            for column_index in range(column_count)
        ]
        if skip_empty_rows and all(value is None or value == "" for value in values):
            continue
        rows.append(
            {column.id: value for column, value in zip(columns, values, strict=True)}
        )

    typed_columns = [
        column.model_copy(
            update={
                "value_type": _inferred_table_value_type(
                    [row[column.id] for row in rows]
                )
            }
        )
        for column in columns
    ]
    return Table(columns=typed_columns, rows=rows)


def _csv_matrix(content: bytes, delimiter: str | None) -> list[list[object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TableFileImportError("CSV files must use UTF-8 encoding") from exc
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(text[:16_384]).delimiter
        except csv.Error:
            delimiter = ","
    return [
        cast(list[object], row)
        for row in csv.reader(StringIO(text), delimiter=delimiter)
    ]


def _xlsx_matrix(
    content: bytes,
    *,
    sheet_name: str | None,
) -> list[list[object]]:
    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise TableFileImportError(
            "The uploaded file is not a readable XLSX workbook"
        ) from exc
    try:
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise TableFileImportError(
                    f"Workbook has no worksheet named {sheet_name!r}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        if worksheet is None:
            raise TableFileImportError("The workbook does not have an active worksheet")
        return [
            [cast(object, value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


@TABLES.node(
    operator_id="table.file.import",
    version=1,
    title="Import table file",
    factory=lambda context: TableFileImportNode(
        uploads_dir=context.uploads_dir,
        unit_of_work=context.uow,
    ),
    staged_upload_inputs=(NodeStagedUploadInput(config_field="uploads"),),
    required_capabilities=(PluginRuntimeCapability.STAGED_UPLOADS,),
    cache_policy=NodeCachePolicy.NEVER,
)
@final
class TableFileImportNode(
    Node[TableFileImportConfig, TableFileImportInput, TableFileImportOutput]
):
    """Import a staged CSV or XLSX file as a table artifact."""

    def __init__(
        self,
        *,
        uploads_dir: Path,
        unit_of_work: StagedUploadUnitOfWorkPort,
    ) -> None:
        self._uploads_dir = uploads_dir.expanduser().resolve()
        self._unit_of_work = unit_of_work

    @override
    async def run(
        self,
        context: NodeExecutionContext,
        config: TableFileImportConfig,
        _inputs: TableFileImportInput,
        /,
    ) -> TableFileImportOutput:
        upload = config.uploads[0]
        try:
            path = await resolve_persisted_staged_upload_path(
                self._uploads_dir,
                self._unit_of_work,
                workspace_id=context.workspace_id,
                upload_key=upload.upload_key,
            )
        except (ValueError, FileNotFoundError) as exc:
            raise TableFileImportError(str(exc)) from exc
        try:
            content = path.read_bytes()
        except OSError as exc:
            raise TableFileImportError(
                f"Failed to read staged table upload {upload.upload_key!r} from {path}"
            ) from exc
        if len(content) != upload.byte_size:
            raise TableFileImportError(
                f"Staged table upload {upload.upload_key!r} changed size: "
                f"expected {upload.byte_size}, got {len(content)}"
            )

        suffix = Path(upload.filename).suffix.casefold()
        if suffix == ".csv":
            matrix = _csv_matrix(content, config.delimiter)
        elif suffix == ".xlsx":
            matrix = _xlsx_matrix(content, sheet_name=config.sheet_name)
        else:
            raise TableFileImportError(
                f"Table upload {upload.filename!r} must end in .csv or .xlsx"
            )
        return TableFileImportOutput(
            table=_table_from_matrix(
                matrix,
                header_row=config.header_row,
                skip_empty_rows=config.skip_empty_rows,
            )
        )


class TableTextNormalizeError(RuntimeError):
    pass


class TableTextNormalizeConfig(NodeConfig):
    source_column: StrictStr = Field(
        min_length=1,
        description="Column id or unique title containing source text.",
    )
    target_column: StrictStr = Field(
        default="normalized_name",
        min_length=1,
        max_length=255,
        description="New column id for normalized text.",
    )
    target_title: StrictStr = Field(
        default="Normalized name",
        max_length=1_024,
        description="Display title of the new column.",
    )
    modernize_historical_cyrillic: StrictBool = True
    transliterate: StrictBool = True
    casefold: StrictBool = True
    remove_diacritics: StrictBool = True
    punctuation_as_spaces: StrictBool = True
    collapse_whitespace: StrictBool = True

    @field_validator("source_column", "target_column")
    @classmethod
    def validate_column_reference(cls, value: str) -> str:
        if value != value.strip():
            raise ValueError("column references must not have surrounding whitespace")
        return value


class TableTextNormalizeInput(NodeInput):
    table: Annotated[
        Table,
        InPort(TABLE_DATA),
        Field(description="Table containing the source text column."),
    ]


class TableTextNormalizeOutput(NodeOutput):
    table: Annotated[
        Table,
        OutPort(TABLE_DATA),
        Field(description="Input table with one added normalized text column."),
    ]


def _resolve_table_column(
    table: Table,
    reference: str,
    *,
    operation: str,
) -> str:
    ids = {column.id for column in table.columns}
    if reference in ids:
        return reference
    title_matches = [column.id for column in table.columns if column.title == reference]
    if len(title_matches) == 1:
        return title_matches[0]
    if len(title_matches) > 1:
        raise ValueError(
            f"{operation} column title {reference!r} is ambiguous; use a column id"
        )
    raise ValueError(f"{operation} table has no column id or title {reference!r}")


_HISTORICAL_CYRILLIC_TRANSLATION = str.maketrans(
    {
        "І": "И",
        "і": "и",
        "Ѣ": "Е",
        "ѣ": "е",
        "Ѳ": "Ф",
        "ѳ": "ф",
        "Ѵ": "И",
        "ѵ": "и",
        "Ѧ": "Я",
        "ѧ": "я",
        "Ѫ": "У",
        "ѫ": "у",
        "Ѡ": "О",
        "ѡ": "о",
        "Ѯ": "Кс",
        "ѯ": "кс",
        "Ѱ": "Пс",
        "ѱ": "пс",
    }
)


def _normalize_table_text(value: str, config: TableTextNormalizeConfig) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    if config.modernize_historical_cyrillic:
        normalized = normalized.translate(_HISTORICAL_CYRILLIC_TRANSLATION)
    if config.transliterate:
        normalized = unidecode(normalized)
    if config.casefold:
        normalized = normalized.casefold()
    if config.remove_diacritics:
        normalized = "".join(
            character
            for character in unicodedata.normalize("NFKD", normalized)
            if not unicodedata.combining(character)
        )
    if config.punctuation_as_spaces:
        normalized = "".join(
            " " if unicodedata.category(character).startswith(("P", "S")) else character
            for character in normalized
        )
    return (
        re.sub(r"\s+", " ", normalized).strip()
        if config.collapse_whitespace
        else normalized.strip()
    )


@TABLES.function_node(
    operator_id="table.text.normalize",
    version=1,
    title="Normalize table text",
    cache_policy=NodeCachePolicy.EXACT,
)
async def normalize_table_text(
    config: TableTextNormalizeConfig,
    inputs: TableTextNormalizeInput,
) -> TableTextNormalizeOutput:
    """Add a configurable normalized-text column to a table."""

    try:
        source_column = _resolve_table_column(
            inputs.table,
            config.source_column,
            operation="Text normalization",
        )
    except ValueError as exc:
        raise TableTextNormalizeError(str(exc)) from exc
    if config.target_column in {column.id for column in inputs.table.columns}:
        raise TableTextNormalizeError(
            f"Target column id {config.target_column!r} already exists"
        )

    rows: list[dict[str, TableValue]] = []
    for row_index, row in enumerate(inputs.table.rows):
        source_value = row[source_column]
        if source_value is not None and not isinstance(source_value, str):
            raise TableTextNormalizeError(
                f"Row {row_index} source column {config.source_column!r} "
                "must contain text or null"
            )
        rows.append(
            {
                **row,
                config.target_column: (
                    None
                    if source_value is None
                    else _normalize_table_text(source_value, config)
                ),
            }
        )
    return TableTextNormalizeOutput(
        table=Table(
            columns=[
                *inputs.table.columns,
                TableColumn(
                    id=config.target_column,
                    title=config.target_title,
                    value_type=TableValueType.TEXT,
                ),
            ],
            rows=rows,
        )
    )


NormalizeTableTextNode = TABLES.nodes[-1].node_class


class FuzzyMatchScorer(StrEnum):
    RATIO = "ratio"
    WEIGHTED_RATIO = "weighted_ratio"
    TOKEN_SORT_RATIO = "token_sort_ratio"


class TableFuzzyMatchError(RuntimeError):
    pass


class TableFuzzyMatchConfig(NodeConfig):
    left_text_column: StrictStr = Field(
        min_length=1,
        description="Left table column id or unique title containing normalized text.",
    )
    left_alias_columns: list[Annotated[StrictStr, Field(min_length=1)]] = Field(
        default_factory=list,
        max_length=8,
        description="Optional additional normalized-text columns on the left.",
    )
    right_text_column: StrictStr = Field(
        min_length=1,
        description="Right table column id or unique title containing normalized text.",
    )
    right_alias_columns: list[Annotated[StrictStr, Field(min_length=1)]] = Field(
        default_factory=list,
        max_length=8,
        description="Optional additional normalized-text columns on the right.",
    )
    left_block_column: StrictStr | None = Field(
        default=None,
        min_length=1,
        description="Optional left context column requiring an exact match.",
    )
    right_block_column: StrictStr | None = Field(
        default=None,
        min_length=1,
        description="Optional right context column requiring an exact match.",
    )
    scorer: FuzzyMatchScorer = FuzzyMatchScorer.WEIGHTED_RATIO
    score_threshold: StrictFloat = Field(default=70.0, ge=0.0, le=100.0)
    max_candidates: StrictInt = Field(default=5, ge=1, le=100)
    include_unmatched: StrictBool = True
    max_comparisons: StrictInt = Field(
        default=2_000_000,
        ge=1,
        le=100_000_000,
        description="Hard ceiling for pairwise string comparisons.",
    )

    @field_validator(
        "left_text_column",
        "right_text_column",
        "left_block_column",
        "right_block_column",
    )
    @classmethod
    def validate_column_reference(cls, value: str | None) -> str | None:
        if value is not None and value != value.strip():
            raise ValueError("column references must not have surrounding whitespace")
        return value

    @field_validator("left_alias_columns", "right_alias_columns")
    @classmethod
    def validate_alias_column_references(cls, values: list[str]) -> list[str]:
        if any(value != value.strip() for value in values):
            raise ValueError("column references must not have surrounding whitespace")
        if len(values) != len(set(values)):
            raise ValueError("alias column references must be unique")
        return values

    @model_validator(mode="after")
    def validate_block_columns(self) -> Self:
        if (self.left_block_column is None) != (self.right_block_column is None):
            raise ValueError(
                "left_block_column and right_block_column must be configured together"
            )
        if self.left_text_column in self.left_alias_columns:
            raise ValueError("left_text_column must not also be a left alias column")
        if self.right_text_column in self.right_alias_columns:
            raise ValueError("right_text_column must not also be a right alias column")
        return self


class TableFuzzyMatchInput(NodeInput):
    left: Annotated[
        Table,
        InPort(TABLE_DATA),
        Field(description="Source records to match."),
    ]
    right: Annotated[
        Table,
        InPort(TABLE_DATA),
        Field(description="Candidate records."),
    ]


class TableFuzzyMatchOutput(NodeOutput):
    matches: Annotated[
        Table,
        OutPort(TABLE_DATA),
        Field(description="Ranked candidate pairs with original source columns."),
    ]


def _fuzzy_score(
    left: str,
    right: str,
    scorer: FuzzyMatchScorer,
) -> float:
    if scorer is FuzzyMatchScorer.RATIO:
        return float(fuzz.ratio(left, right))
    if scorer is FuzzyMatchScorer.TOKEN_SORT_RATIO:
        return float(fuzz.token_sort_ratio(left, right))
    return float(fuzz.WRatio(left, right))


def _blocking_value(
    value: TableValue,
    *,
    row_index: int,
    column: str,
    side: str,
) -> tuple[type[object], object]:
    if isinstance(value, list | dict):
        raise TableFuzzyMatchError(
            f"{side} row {row_index} blocking column {column!r} "
            "must contain a scalar value"
        )
    return type(value), cast(object, value)


def _prefixed_match_column_id(
    side: Literal["left", "right"],
    column: TableColumn,
    column_index: int,
) -> str:
    candidate = f"{side}__{column.id}"
    if len(candidate) <= 255:
        return candidate
    digest = sha256(column.id.encode("utf-8")).hexdigest()[:12]
    return f"{side}__column_{column_index + 1}_{digest}"


@TABLES.function_node(
    operator_id="table.fuzzy_match",
    version=1,
    title="Fuzzy match tables",
    cache_policy=NodeCachePolicy.EXACT,
)
async def fuzzy_match_tables(
    config: TableFuzzyMatchConfig,
    inputs: TableFuzzyMatchInput,
) -> TableFuzzyMatchOutput:
    """Rank fuzzy text matches between rows from two tables."""

    try:
        left_text_columns = [
            _resolve_table_column(
                inputs.left,
                reference,
                operation="Fuzzy match left",
            )
            for reference in [
                config.left_text_column,
                *config.left_alias_columns,
            ]
        ]
        right_text_columns = [
            _resolve_table_column(
                inputs.right,
                reference,
                operation="Fuzzy match right",
            )
            for reference in [
                config.right_text_column,
                *config.right_alias_columns,
            ]
        ]
        if len(left_text_columns) != len(set(left_text_columns)):
            raise ValueError("Fuzzy match left text columns must resolve uniquely")
        if len(right_text_columns) != len(set(right_text_columns)):
            raise ValueError("Fuzzy match right text columns must resolve uniquely")
        left_block_column = (
            _resolve_table_column(
                inputs.left,
                config.left_block_column,
                operation="Fuzzy match left block",
            )
            if config.left_block_column is not None
            else None
        )
        right_block_column = (
            _resolve_table_column(
                inputs.right,
                config.right_block_column,
                operation="Fuzzy match right block",
            )
            if config.right_block_column is not None
            else None
        )
    except ValueError as exc:
        raise TableFuzzyMatchError(str(exc)) from exc

    right_text_values: list[list[tuple[str, str]]] = []
    right_rows_by_block: dict[tuple[type[object], object] | None, list[int]] = {}
    for right_index, row in enumerate(inputs.right.rows):
        row_text_values: list[tuple[str, str]] = []
        for configured_column, resolved_column in zip(
            [config.right_text_column, *config.right_alias_columns],
            right_text_columns,
            strict=True,
        ):
            value = row[resolved_column]
            if value is not None and not isinstance(value, str):
                raise TableFuzzyMatchError(
                    f"Right row {right_index} text column "
                    f"{configured_column!r} must contain text or null"
                )
            if value is not None:
                row_text_values.append((resolved_column, value))
        right_text_values.append(row_text_values)
        block = (
            _blocking_value(
                row[right_block_column],
                row_index=right_index,
                column=config.right_block_column or right_block_column,
                side="Right",
            )
            if right_block_column is not None
            else None
        )
        right_rows_by_block.setdefault(block, []).append(right_index)

    comparison_count = 0
    ranked_matches: list[
        tuple[
            int,
            int | None,
            float | None,
            int | None,
            str | None,
            str | None,
        ]
    ] = []
    for left_index, left_row in enumerate(inputs.left.rows):
        left_text_values: list[tuple[str, str]] = []
        for configured_column, resolved_column in zip(
            [config.left_text_column, *config.left_alias_columns],
            left_text_columns,
            strict=True,
        ):
            value = left_row[resolved_column]
            if value is not None and not isinstance(value, str):
                raise TableFuzzyMatchError(
                    f"Left row {left_index} text column "
                    f"{configured_column!r} must contain text or null"
                )
            if value is not None:
                left_text_values.append((resolved_column, value))
        block = (
            _blocking_value(
                left_row[left_block_column],
                row_index=left_index,
                column=config.left_block_column or left_block_column,
                side="Left",
            )
            if left_block_column is not None
            else None
        )
        right_indices = right_rows_by_block.get(block, [])
        scored: list[tuple[float, int, str, str]] = []
        if left_text_values:
            for right_index in right_indices:
                candidate_text_values = right_text_values[right_index]
                if not candidate_text_values:
                    continue
                comparison_count += len(left_text_values) * len(candidate_text_values)
                if comparison_count > config.max_comparisons:
                    raise TableFuzzyMatchError(
                        "Fuzzy matching would exceed max_comparisons "
                        f"({comparison_count} > {config.max_comparisons}); configure "
                        "blocking columns or increase the explicit ceiling"
                    )
                best_match: tuple[float, str, str] | None = None
                for left_column, left_text_value in left_text_values:
                    for right_column, right_text_value in candidate_text_values:
                        score = _fuzzy_score(
                            left_text_value,
                            right_text_value,
                            config.scorer,
                        )
                        if best_match is None or score > best_match[0]:
                            best_match = (score, left_column, right_column)
                if best_match is not None and best_match[0] >= config.score_threshold:
                    scored.append(
                        (
                            best_match[0],
                            right_index,
                            best_match[1],
                            best_match[2],
                        )
                    )
        scored.sort(key=lambda item: (-item[0], item[1]))
        selected = scored[: config.max_candidates]
        if selected:
            ranked_matches.extend(
                (
                    left_index,
                    right_index,
                    score,
                    rank,
                    left_column,
                    right_column,
                )
                for rank, (
                    score,
                    right_index,
                    left_column,
                    right_column,
                ) in enumerate(selected, start=1)
            )
        elif config.include_unmatched:
            ranked_matches.append((left_index, None, None, None, None, None))

    left_output_ids = [
        _prefixed_match_column_id("left", column, index)
        for index, column in enumerate(inputs.left.columns)
    ]
    right_output_ids = [
        _prefixed_match_column_id("right", column, index)
        for index, column in enumerate(inputs.right.columns)
    ]
    output_columns = [
        TableColumn(
            id="left_row_index",
            title="Source row index",
            value_type=TableValueType.INTEGER,
        ),
        TableColumn(
            id="right_row_index",
            title="Candidate row index",
            value_type=TableValueType.INTEGER,
        ),
        TableColumn(
            id="match_score",
            title="Match score",
            value_type=TableValueType.NUMBER,
        ),
        TableColumn(
            id="match_rank",
            title="Candidate rank",
            value_type=TableValueType.INTEGER,
        ),
        TableColumn(
            id="match_scorer",
            title="Match scorer",
            value_type=TableValueType.TEXT,
        ),
        TableColumn(
            id="match_left_column",
            title="Matched source column",
            value_type=TableValueType.TEXT,
        ),
        TableColumn(
            id="match_right_column",
            title="Matched candidate column",
            value_type=TableValueType.TEXT,
        ),
        *[
            TableColumn(
                id=output_id,
                title=f"Source · {column.title or column.id}",
                value_type=column.value_type,
            )
            for output_id, column in zip(
                left_output_ids,
                inputs.left.columns,
                strict=True,
            )
        ],
        *[
            TableColumn(
                id=output_id,
                title=f"Candidate · {column.title or column.id}",
                value_type=column.value_type,
            )
            for output_id, column in zip(
                right_output_ids,
                inputs.right.columns,
                strict=True,
            )
        ],
    ]
    output_rows: list[dict[str, TableValue]] = []
    for (
        left_index,
        right_index,
        score,
        rank,
        match_left_column,
        match_right_column,
    ) in ranked_matches:
        left_row = inputs.left.rows[left_index]
        right_row = inputs.right.rows[right_index] if right_index is not None else None
        output_row: dict[str, TableValue] = {
            "left_row_index": left_index,
            "right_row_index": right_index,
            "match_score": score,
            "match_rank": rank,
            "match_scorer": config.scorer.value,
            "match_left_column": match_left_column,
            "match_right_column": match_right_column,
        }
        output_row.update(
            {
                output_id: left_row[column.id]
                for output_id, column in zip(
                    left_output_ids,
                    inputs.left.columns,
                    strict=True,
                )
            }
        )
        output_row.update(
            {
                output_id: (right_row[column.id] if right_row is not None else None)
                for output_id, column in zip(
                    right_output_ids,
                    inputs.right.columns,
                    strict=True,
                )
            }
        )
        output_rows.append(output_row)

    return TableFuzzyMatchOutput(
        matches=Table(columns=output_columns, rows=output_rows)
    )


FuzzyMatchTablesNode = TABLES.nodes[-1].node_class


__all__ = [
    "FuzzyMatchTablesNode",
    "FuzzyMatchScorer",
    "NormalizeTableTextNode",
    "TableFileImportConfig",
    "TableFileImportError",
    "TableFileImportInput",
    "TableFileImportNode",
    "TableFileImportOutput",
    "TableFileUploadItem",
    "TableFuzzyMatchConfig",
    "TableFuzzyMatchError",
    "TableFuzzyMatchInput",
    "TableFuzzyMatchOutput",
    "TableTextNormalizeConfig",
    "TableTextNormalizeError",
    "TableTextNormalizeInput",
    "TableTextNormalizeOutput",
    "fuzzy_match_tables",
    "normalize_table_text",
]
